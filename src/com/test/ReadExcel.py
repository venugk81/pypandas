from openpyxl import load_workbook
workbook = load_workbook("E:\python projects\data\excel_data.xlsx")
saveAsFile ="E:\python projects\data\excel_data1.xlsx"
sheet_name = workbook["Sheet2"]

max_row = sheet_name.max_row
max_column = sheet_name.max_column
print(max_column)

excel_values = []
headers = []
for j in range(1, max_column):
    headers.append(sheet_name.cell(1, column=j).value)
print("Headers: ", headers)
print("header values------------------")
for i in range(len(headers)):
    print(headers[i])

print("headers-----------------------------------------")

for i in range(2, max_row + 1):
    dict={}
    # iterate over all coluns
    for j in range(1, max_column + 1):
        # get particular cell value
        cell_obj = sheet_name.cell(row=i, column=j)

        # print cell value
        if j == 8:
            cell_obj.value = "PASS"


        print(cell_obj.value, end=' | ')
    # print new line
    print('\n')
    excel_values.append(dict)

workbook.save(saveAsFile)

print("========end==========")
print(excel_values)